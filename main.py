from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver

import win32com.client

import openpyxl

##IL 150/335-45/4 конец

Excel = win32com.client.Dispatch("Excel.Application")
from selenium.webdriver.common.keys import Keys


def Exel_format(filename, NamePump, ArtPump, PricePump, Speed_of_rotation, Rated_power, Current, Power_Factor,
                Efficiency_50, Efficiency_75, Efficiency_100):
    wb_in = Excel.Workbooks.Open(filename)
    sheet = wb_in.ActiveSheet


    # получаем значения цепочки A1:A2
    vals1 = [r[0].value for r in sheet.Range("B5:B100")]
    vals2 = [r[0].value for r in sheet.Range("C5:C100")]
    vals3 = [r[0].value for r in sheet.Range("D5:D100")]
    vals4 = [r[0].value for r in sheet.Range("E5:E100")]
    vals5 = [r[0].value for r in sheet.Range("F5:F100")]

    #поиск максимального кпд
    buf = vals4
    vals4=[]
    for i in buf:
        if i != None:
            vals4.append(i)
    print(vals4)
    eta_bep = max(vals4)
    Q_bep=vals1[vals4.index(max(vals4))]

    eta_bep=float("{0:.2f}".format(eta_bep))
    Q_bep=float("{0:.2f}".format(Q_bep))
    # # очистка поля
    # for i in range(1, 10):
    #     for j in range(1, 100):
    #         sheet.Cells(j, i).value = ""
    # сохраняем рабочую книгу
    wb_in.Save()

    # закрываем ее
    wb_in.Close()

    # закрываем COM объект
    Excel.Quit()

##создаем новый файл exel

    #wb.create_sheet(title='Sheet', index=0)
    #sheet = wb['Sheet']

    ##Закрашиваем ячейки

    wb = openpyxl.Workbook()

    for i in range(1, 12):
        cell = wb.active.cell(column=i, row=7)
        cell.fill = openpyxl.styles.PatternFill(start_color='7ffa84', end_color='7ffa84', fill_type='solid')
    for i in range(13, 20):
        cell = wb.active.cell(column=i, row=7)
        cell.fill = openpyxl.styles.PatternFill(start_color='7ff2fa', end_color='7ff2fa', fill_type='solid')

    NamePump= "CronoLine-"+NamePump
    NamePumpFile = NamePump.replace('/','_')
    wb.save(u'D:\\gomno\\рабочий стол старая ос\\учеба\\7 семестр\\шабашка\\Итог\\'+NamePumpFile+'.xlsx')

###Открываем новый файл
    wb = Excel.Workbooks.Open(
        u'D:\\gomno\\рабочий стол старая ос\\учеба\\7 семестр\\шабашка\\Итог\\'+NamePumpFile+'.xlsx')
    sheet = wb.ActiveSheet


###Удаляем нежелателльные символы
    Speed_of_rotation = Speed_of_rotation.replace("1/min", "").replace(" ", "")
    Rated_power = Rated_power.replace("kW", "").replace(" ", "")
    Current = Current.replace("A", "").replace(" ", "")
    Efficiency_50 = Efficiency_50.replace("%", "").replace(" ", "")
    Efficiency_75 = Efficiency_75.replace("%", "").replace(" ", "")
    Efficiency_100 = Efficiency_100.replace("%", "").replace(" ", "")
###

    # записываем значение таблиц в определенную ячейку
    sheet.Cells(1, 1).value = "Производитель"
    sheet.Cells(1, 4).value = "Wilo"

    sheet.Cells(2, 1).value = "Название"
    sheet.Cells(2, 4).value = NamePump

    sheet.Cells(3, 1).value = "Артикул"
    sheet.Cells(3, 4).value = ArtPump

    sheet.Cells(4, 1).value = "Цена"
    sheet.Cells(4, 4).value = PricePump

    sheet.Cells(6, 1).value = "Насос"
    sheet.Cells(6, 13).value = "Двигатель"

    sheet.Cells(7, 1).value = "Q, m³/h"
    sheet.Cells(7, 2).value = "H, m"

    sheet.Cells(7, 3).value = "NPSH, m"

    sheet.Cells(7, 4).value = "η, %"
    sheet.Cells(7, 5).value = "P₂, kW"

    # sheet.Cells(7, 7).value = "H_bep"
    sheet.Cells(7, 7).value = "Q_bep"
    sheet.Cells(7, 8).value = "eta_bep"

    sheet.Cells(7, 10).value = "Q_min"
    sheet.Cells(7, 11).value = "Q_max"

    sheet.Cells(7, 13).value = "Частота вращ."
    sheet.Cells(8, 13).value = Speed_of_rotation

    sheet.Cells(7, 14).value = "Ном.мощность"
    sheet.Cells(8, 14).value = Rated_power

    sheet.Cells(7, 15).value = "Ток"
    sheet.Cells(8, 15).value = Current

    sheet.Cells(7, 16).value = "Коэф.мощн."
    sheet.Cells(8, 16).value = Power_Factor

    sheet.Cells(7, 17).value = "КПД 50%"
    sheet.Cells(8, 17).value = Efficiency_50

    sheet.Cells(7, 18).value = "КПД 75%"
    sheet.Cells(8, 18).value = Efficiency_75

    sheet.Cells(7, 19).value = "КПД 100%"
    sheet.Cells(8, 19).value = Efficiency_100
    # Current,Power_Factor,Efficiency_50,Efficiency_75,Efficiency_100
    print("Введите Q_bep")
    sheet.Cells(8, 7).value = Q_bep
    print(Q_bep)
    print("Введите eta_bep")
    sheet.Cells(8, 8).value = eta_bep
    print(eta_bep)
    print("Введите Q_min")
    sheet.Cells(8, 10).value = input()
    print("Введите Q_max")
    sheet.Cells(8, 11).value = input()
    # записываем последовательность
    i = 8
    for rec in vals1:
        sheet.Cells(i, 1).value = rec
        i += 1

    i = 8
    for rec in vals2:
        sheet.Cells(i, 2).value = rec
        i += 1


    i = 8
    for rec in vals3:
        sheet.Cells(i, 3).value = rec
        i += 1

    i = 8
    for rec in vals4:
        sheet.Cells(i, 4).value = rec
        i += 1

    i = 8
    for rec in vals5:
        sheet.Cells(i, 5).value = rec
        i += 1

# сохраняем рабочую книгу
    wb.Save()
    # закрываем ее
    wb.Close()
    # закрываем COM объект
    Excel.Quit()


def lasf_file():
    import os
    path = u'D:\\gomno\\рабочий стол старая ос\\учеба\\7 семестр\\шабашка\\Итог\\'
    files = os.listdir(path)
    files = [os.path.join(path, file) for file in files]
    files = [file for file in files if os.path.isfile(file)]
    name=max(files, key=os.path.getmtime)
    print(name)
    return name

driver = webdriver.Chrome(
    'D:\\gomno\\рабочий стол старая ос\\учеба\\7 семестр\\шабашка\\chromedriver_win32 (1)\\chromedriver.exe')
driver.maximize_window()
wait = WebDriverWait(driver, 20)

driver.get("https://www.wilo-select.com/PumpProdConfig.aspx")
sleep(3)

driver.find_element_by_xpath(".//input[@name='loginName']").send_keys("gitvanya34")
driver.find_element_by_xpath(".//input[@name='loginPassword']").send_keys("6Cr-Sc8-Lir-inc")
driver.find_element_by_xpath(".//button[@name='ctl34']").click()

sleep(10)
wait.until(EC.element_to_be_clickable((By.XPATH, ".//button[@name='ctl149']")))
driver.find_element_by_xpath(".//button[@name='ctl149']").click()
wait.until(EC.element_to_be_clickable((By.XPATH, ".//input[@name='pnlBar_SelectionType$i0$i0$ddSeries']")))
driver.find_element_by_xpath(".//input[@name='pnlBar_SelectionType$i0$i0$ddSeries']").click()
driver.find_element_by_xpath(".//input[@name='pnlBar_SelectionType$i0$i0$ddSeries']").send_keys("CronoLine-IL")
driver.find_element_by_xpath(".//*[@class='rpText']").click()

sleep(10)
print("с какого элемента начать?")
start_num=(int)(input())
if start_num==1:
    wait.until(EC.element_to_be_clickable(
        (By.XPATH, ".//span[@class='rmText rmExpandDown' and contains(text(),'Экспорт')]/parent::a/parent::li")))
    driver.find_element_by_xpath(
        ".//span[@class='rmText rmExpandDown' and contains(text(),'Экспорт')]/parent::a/parent::li").click()
    driver.find_element_by_xpath(".//*[contains(text(),'Экспортировать точки кривой')]").click()
    # редактирование файла

    ##
    NamePump = driver.find_element_by_xpath(
        ".//tr[ contains(@class, 'rgSelectedRow')] //div[@ class ='gridValueDivInner PRODDESC']").get_attribute("title")
    ArtPump = driver.find_element_by_xpath(
        ".//tr[ contains(@class, 'rgSelectedRow')]//div[@ class ='gridValueDivInner ARTNR']").get_attribute("title")
    PricePump = driver.find_element_by_xpath(
        ".//tr[ contains(@class, 'rgSelectedRow')]//div[@class='gridValueDivInner FITTPRICE' ]").get_attribute("title")
    print(NamePump)
    print(ArtPump)
    print(PricePump)
    ##
    ###
    infoRow_list = driver.find_elements_by_xpath(".//a[@class='mpMenuItemLbl']")  # элементы списка "Информация"
    infoRow_list[1].click()  # выбор жлемента "Описнаие продукта"

    wait.until(EC.element_to_be_clickable((By.XPATH,
                                           ".//span[contains(text(),'Номинальная частота вращения')]/following-sibling::span")))  # ждем появления списка

    Speed_of_rotation = driver.find_element_by_xpath(
        ".//span[contains(text(),'Номинальная частота вращения')]/following-sibling::span").text
    Rated_power = driver.find_element_by_xpath(
        ".//span[contains(text(),'Номинальная мощность')]/following-sibling::span").text
    Current = driver.find_element_by_xpath(".//span[contains(text(),'Номинальный ток')]/following-sibling::span").text
    Power_Factor = driver.find_element_by_xpath(
        ".//span[contains(text(),'Коэффициент мощности')]/following-sibling::span").text
    Efficiency_50 = driver.find_element_by_xpath(".//span[contains(text(),'ηm 50')]/following-sibling::span").text
    Efficiency_75 = driver.find_element_by_xpath(".//span[contains(text(),'ηm 75')]/following-sibling::span").text
    Efficiency_100 = driver.find_element_by_xpath(".//span[contains(text(),'ηm 100')]/following-sibling::span").text

    print(Speed_of_rotation)
    print(Rated_power)
    print(Current)
    print(Power_Factor)
    print(Efficiency_50)
    print(Efficiency_75)
    print(Efficiency_100)

    infoRow_list = driver.find_elements_by_xpath(".//a[@class='mpMenuItemLbl']")  # элементы списка "Информация"
    infoRow_list[0].click()  # выбор жлемента "Кривые насосов"

    ###
    print("введите название файла")
    filename = lasf_file()
    Exel_format(filename, NamePump, ArtPump, PricePump, Speed_of_rotation, Rated_power, Current, Power_Factor,
                Efficiency_50, Efficiency_75, Efficiency_100)
# переход на следующую строку
pump_list = driver.find_elements_by_xpath(".//*[@class='rgRow rgEven' or @class='rgRow rgOdd']")

count_element=1
for pump in pump_list:
    count_element+=1
    if count_element >= start_num:
    ##Норм работающий блок
        pump.click()

        wait.until(EC.element_to_be_clickable(
            (By.XPATH, ".//span[@class='rmText rmExpandDown' and contains(text(),'Экспорт')]/parent::a/parent::li")))
        sleep(5)
        driver.find_element_by_xpath(
            ".//span[@class='rmText rmExpandDown' and contains(text(),'Экспорт')]/parent::a/parent::li").click()

        Export_list = driver.find_elements_by_xpath(".//*[contains(text(),'Экспортировать точки кривой')]")


        Export_list[-1].click()  # при выборе седующего жлемента списка образуются новые элементы

        ##

        NamePump = driver.find_element_by_xpath(
            ".//tr[ contains(@class, 'rgSelectedRow')] //div[@ class ='gridValueDivInner PRODDESC']").get_attribute("title")
        ArtPump = driver.find_element_by_xpath(
            ".//tr[ contains(@class, 'rgSelectedRow')]//div[@ class ='gridValueDivInner ARTNR']").get_attribute("title")
        PricePump = driver.find_element_by_xpath(
            ".//tr[ contains(@class, 'rgSelectedRow')]//div[@class='gridValueDivInner FITTPRICE' ]").get_attribute("title")

        print(NamePump)
        print(ArtPump)
        print(PricePump)

        infoRow_list = driver.find_elements_by_xpath(".//a[@class='mpMenuItemLbl']")  # элементы списка "Информация"
        infoRow_list[1].click()  # выбор жлемента "Описнаие продукта"

        sleep(5)
        # wait.until(EC.elementIsVisible((By.XPATH,".//span[contains(text(),'Номинальная частота вращения')]/following-sibling::span")))  # ждем появления списка
        Speed_of_rotation = driver.find_element_by_xpath(
            ".//span[contains(text(),'Номинальная частота вращения')]/following-sibling::span").text

        # wait.until(EC.ElementIsVisible((By.XPATH, ".//span[contains(text(),'Номинальная мощность')]/following-sibling::span")))
        Rated_power = driver.find_element_by_xpath(
            ".//span[contains(text(),'Номинальная мощность')]/following-sibling::span").text

        # wait.until(EC.ElementIsVisible((By.XPATH, ".//span[contains(text(),'Номинальный ток')]/following-sibling::span")))
        Current = driver.find_element_by_xpath(".//span[contains(text(),'Номинальный ток')]/following-sibling::span").text

        # wait.until(EC.ElementIsVisible((By.XPATH, ".//span[contains(text(),'Коэффициент мощности')]/following-sibling::span")))
        Power_Factor = driver.find_element_by_xpath(
            ".//span[contains(text(),'Коэффициент мощности')]/following-sibling::span").text

        # wait.until(EC.ElementIsVisible((By.XPATH, ".//span[contains(text(),'ηm 50')]/following-sibling::span")))
        Efficiency_50 = driver.find_element_by_xpath(".//span[contains(text(),'ηm 50')]/following-sibling::span").text

        # wait.until(EC.ElementIsVisible((By.XPATH, ".//span[contains(text(),'ηm 75')]/following-sibling::span")))
        Efficiency_75 = driver.find_element_by_xpath(".//span[contains(text(),'ηm 75')]/following-sibling::span").text

        #  wait.until(EC.ElementIsVisible((By.XPATH, ".//span[contains(text(),'ηm 100')]/following-sibling::span")))
        Efficiency_100 = driver.find_element_by_xpath(".//span[contains(text(),'ηm 100')]/following-sibling::span").text

        print(Speed_of_rotation)
        print(Rated_power)
        print(Current)
        print(Power_Factor)
        print(Efficiency_50)
        print(Efficiency_75)
        print(Efficiency_100)

        infoRow_list = driver.find_elements_by_xpath(".//a[@class='mpMenuItemLbl']")  # элементы списка "Информация"
        infoRow_list[0].click()  # выбор жлемента "Кривые насосов"

        ##
        print("введите название файла")

        filename = lasf_file()
        # print(filename)

        Exel_format(filename, NamePump, ArtPump, PricePump, Speed_of_rotation, Rated_power, Current, Power_Factor,
                    Efficiency_50, Efficiency_75, Efficiency_100)
    ##

##TODO 1)закинуть данные в функцию экселя,+
# TODO 2)сделать ввод с консоли данных с графиков+
# TODO 3) Настроить папку загрузок и папку взятия+
# TODO 4) Удалить лишнии листы
# TODO 5) Переименовать файлы


##
# //div[text()=’Тема’]/parent::div
# 6Cr-Sc8-Lir-inc
# gitvanya34
# .//input[@name='loginName']
# .//input[@name='loginPassword']
# driver.close()
